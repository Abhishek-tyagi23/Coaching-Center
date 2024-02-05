import openpyxl

def create_student_form():
    # Create a new Excel workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Add headers to the sheet
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Father_Name'
    sheet['C1'] = 'Age'
    sheet['D1'] = 'Mobile_number'
    sheet['D1'] = 'Course'

    # Get student information
    name = input("Enter student's name: ")
    Father_Name = input("Enter student's Father name: ")
    age = input("Enter student's age: ")
    Mobile_number = input("Enter student's Mobile_number: ")
    Course = input("Enter Course Name: ")
    

    # Add student information to the sheet
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=Father_Name)
    sheet.cell(row=row, column=3, value=age)
    sheet.cell(row=row, column=4, value=Mobile_number)
    sheet.cell(row=row, column=4, value=Course)

    # Save the workbook
    wb.save('student.xlsx')

    print(f"You Form is Successful Submitted")

if __name__ == "__main__":
    create_student_form()